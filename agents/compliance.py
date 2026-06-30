import streamlit as st
import base64
import json
from datetime import date
from utils.database import insert_bid_analysis, get_bid_analyses, delete_bid_analysis
from utils.openai_helper import openai_client

COMPLIANCE_PROMPT = """You are an expert tender compliance, technical evaluation, procurement, and contract review assistant.
You will receive up to 10 uploaded files related to one bid. The files may include RFQ/tender documents, BOQ, drawings, scope of work, terms and conditions, supplier quotations, and the final client quotation prepared by the user.
Your task is to analyze all uploaded documents and verify whether the client quotation is technically, commercially, and contractually compliant with the RFQ/tender requirements and supplier quotes.

Perform the analysis in the following structure:

**1. Bid Summary**
Summarize the bid title, customer name, RFQ scope, key deliverables, important dates if available, and overall purpose of the quotation.

**2. Document Classification**
Classify each uploaded file as one of: RFQ/Tender Document, BOQ/Scope of Work, Technical Specification, Supplier Quote, Client Quote, Commercial Terms, Contract Terms, Other.

**3. RFQ Requirement Extraction**
Extract all important RFQ requirements including: Technical scope, BOQ line items, Quantities, Mandatory specifications, Brand/model requirements, Delivery requirements, Warranty requirements, Installation/testing/commissioning requirements, Documentation requirements, Training requirements, SLA/support requirements, Compliance/certification requirements, Payment terms, Penalty/liquidated damages clauses, Validity requirements, Submission requirements, Any exclusions or special conditions.

**4. Supplier Quote Comparison**
Compare supplier quotes against RFQ requirements and identify: Items covered, Items missing, Quantity mismatches, Specification mismatches, Brand/model mismatches, Warranty differences, Delivery differences, Commercial differences, Assumptions or exclusions by supplier, Risks passed from supplier to user.

**5. Client Quote Compliance Check**
Compare the client quotation against the RFQ and supplier quotes. Check whether the client quote includes all required items, quantities, services, warranties, terms, delivery commitments, and scope obligations.

**6. Compliance Matrix**
Create a table with columns: RFQ Requirement | RFQ Reference/Source | Supplier Quote Coverage | Client Quote Coverage | Status (Compliant/Partially Compliant/Non-Compliant/Not Mentioned) | Risk Level (Low/Medium/High) | Recommended Action

**7. Missing Items and Gaps**
List all missing, unclear, underquoted, or non-compliant items separated into: Technical gaps, Commercial gaps, Contractual gaps, BOQ/quantity gaps, Delivery/warranty/support gaps.

**8. Risk Analysis**
Identify risks before submitting the quote including: Financial risk, Scope creep risk, Penalty risk, Delivery risk, Supplier dependency risk, Warranty/support risk, Compliance risk, Contractual liability risk.

**9. Recommended Corrections Before Submission**
Provide clear corrections that should be made to the client quotation before submission. Include wording where useful.

**10. Suggested Exclusions and Assumptions**
Suggest suitable exclusions, assumptions, and clarifications to protect the bidder.

**11. Final Decision**
Give one of: Ready to Submit / Submit After Minor Corrections / Submit After Major Corrections / Do Not Submit Until Clarified. Explain the reason clearly.

Important rules:
- Do not assume missing information as compliant.
- If a requirement is not found in the client quote, mark it as Not Mentioned or Non-Compliant.
- If the supplier quote excludes something required by the RFQ, highlight it clearly.
- If there is conflict between RFQ, supplier quote, and client quote, treat RFQ as the primary requirement.
- Use file names and section references wherever possible.
- Do not invent technical specifications, prices, quantities, delivery dates, or contract terms.
- If information is unclear, state that clarification is required.
- Output should be professional, concise, and suitable for bid review before submission."""


def extract_file_content(uploaded_file):
    file_name = uploaded_file.name
    file_bytes = uploaded_file.read()
    ext = file_name.lower().split(".")[-1]

    if ext in ["jpg", "jpeg", "png", "gif", "bmp", "webp"]:
        b64 = base64.b64encode(file_bytes).decode("utf-8")
        mime = "image/jpeg" if ext in ["jpg", "jpeg"] else f"image/{ext}"
        return {"type": "image", "name": file_name, "b64": b64, "mime": mime}

    elif ext == "pdf":
        try:
            import pdfplumber
            import io
            text = ""
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t:
                        text += t + "\n"
            return {"type": "text", "name": file_name, "content": text[:15000]}
        except Exception as e:
            return {"type": "text", "name": file_name, "content": f"[PDF extraction failed: {e}]"}

    elif ext in ["doc", "docx"]:
        try:
            import zipfile
            import io
            import re
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
                with z.open("word/document.xml") as xml_file:
                    xml_content = xml_file.read().decode("utf-8")
            # Remove XML tags and extract plain text
            text = re.sub(r'<w:p[ >]', '\n', xml_content)
            text = re.sub(r'<w:tr[ >]', '\n', text)
            text = re.sub(r'<w:tc[ >]', ' | ', text)
            text = re.sub(r'<[^>]+>', '', text)
            text = re.sub(r'\n\s*\n', '\n', text).strip()
            return {"type": "text", "name": file_name, "content": text[:15000]}
        except Exception as e:
            return {"type": "text", "name": file_name, "content": f"[Word extraction failed: {e}]"}

    elif ext in ["xls", "xlsx", "csv"]:
        try:
            import openpyxl
            import io
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            text = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text += f"\n[Sheet: {sheet}]\n"
                for row in ws.iter_rows(values_only=True):
                    row_text = "\t".join([str(c) if c is not None else "" for c in row])
                    if row_text.strip():
                        text += row_text + "\n"
            return {"type": "text", "name": file_name, "content": text[:15000]}
        except Exception as e:
            return {"type": "text", "name": file_name, "content": f"[Excel extraction failed: {e}]"}

    elif ext == "txt":
        try:
            text = file_bytes.decode("utf-8", errors="replace")
            return {"type": "text", "name": file_name, "content": text[:15000]}
        except Exception as e:
            return {"type": "text", "name": file_name, "content": f"[Text extraction failed: {e}]"}

    else:
        return {"type": "text", "name": file_name, "content": f"[Unsupported file type: {ext}]"}


def build_messages(bid_title, customer, files_content):
    content = []
    content.append({"type": "text", "text": COMPLIANCE_PROMPT})
    content.append({"type": "text", "text": f"\n\nBid Title: {bid_title}\nCustomer: {customer}\n\nUploaded files:\n"})

    for fc in files_content:
        if fc["type"] == "image":
            content.append({"type": "text", "text": f"\n[File: {fc['name']} — Image]"})
            content.append({"type": "image_url", "image_url": {"url": f"data:{fc['mime']};base64,{fc['b64']}"}})
        else:
            content.append({"type": "text", "text": f"\n[File: {fc['name']}]\n{fc['content']}\n"})

    return [{"role": "user", "content": content}]


def show_compliance_agent(org_id, user_email, user_role):
    tabs = st.tabs(["📋 New Analysis", "📂 Saved Analyses"])
    tab1, tab2 = tabs[0], tabs[1]

    # ── TAB 1: NEW ANALYSIS ──
    with tab1:
        col_head, col_new = st.columns([4, 1])
        with col_head:
            st.subheader("📋 New Compliance Bid Analysis")
        with col_new:
            st.write("")
            if st.button("🔄 New Analysis", use_container_width=True):
                for k in ["compliance_result", "compliance_title", "compliance_customer", "extracted_files", "prev_file_count"]:
                    st.session_state.pop(k, None)
                st.session_state["upload_key"] = st.session_state.get("upload_key", 0) + 1
                st.rerun()

        upload_key = st.session_state.get("upload_key", 0)

        col1, col2 = st.columns(2)
        with col1:
            bid_title = st.text_input("📌 Bid Title", placeholder="e.g. Network Equipment Supply – Tender 2025", key=f"bid_title_{upload_key}")
        with col2:
            customer = st.text_input("🏢 Customer Name", placeholder="e.g. Ministry of Health", key=f"customer_{upload_key}")

        st.markdown("---")
        st.markdown("**📎 Upload Documents** *(Max 10 files — PDF, Word, Excel, Images)*")

        col_rfq, col_sup, col_client = st.columns(3)
        with col_rfq:
            st.caption("📄 RFQ / Tender Documents")
            rfq_files = st.file_uploader(
                "RFQ files", type=["pdf","doc","docx","xls","xlsx","jpg","jpeg","png"],
                accept_multiple_files=True, key=f"rfq_files_{upload_key}", label_visibility="collapsed"
            )
        with col_sup:
            st.caption("🏭 Supplier Quotes")
            supplier_files = st.file_uploader(
                "Supplier files", type=["pdf","doc","docx","xls","xlsx","jpg","jpeg","png"],
                accept_multiple_files=True, key=f"supplier_files_{upload_key}", label_visibility="collapsed"
            )
        with col_client:
            st.caption("📤 Client Quotation")
            client_files = st.file_uploader(
                "Client files", type=["pdf","doc","docx","xls","xlsx","jpg","jpeg","png"],
                accept_multiple_files=True, key=f"client_files_{upload_key}", label_visibility="collapsed"
            )

        all_files = list(rfq_files or []) + list(supplier_files or []) + list(client_files or [])
        total = len(all_files)

        prev_count = st.session_state.get("prev_file_count", 0)
        if total != prev_count:
            st.session_state.pop("extracted_files", None)
            st.session_state["prev_file_count"] = total

        if total > 0:
            st.info(f"📎 **{total} file(s) attached** — {', '.join([f.name for f in all_files])}")

        if total > 10:
            st.error(f"⚠️ You have attached **{total} files**. Maximum allowed is **10**. Please remove {total - 10} file(s) before analysing.")
            st.stop()

        st.markdown("---")
        analyze_btn = st.button(
            "🔍 ANALYZE", type="primary", use_container_width=True,
            disabled=(total == 0 or not bid_title or not customer)
        )

        if not bid_title or not customer:
            st.caption("⚠️ Please enter Bid Title and Customer Name to enable analysis.")

        if total > 0 and "extracted_files" not in st.session_state:
            files_content = []
            for f in all_files:
                fc = extract_file_content(f)
                files_content.append(fc)
            st.session_state["extracted_files"] = files_content


        if analyze_btn and bid_title and customer and total > 0:
            files_content = st.session_state.get("extracted_files", [])
            if not files_content:
                files_content = []
                for f in all_files:
                    fc = extract_file_content(f)
                    files_content.append(fc)

            with st.spinner("🤖 AI is analysing your bid for compliance... This may take a minute..."):
                messages = build_messages(bid_title, customer, files_content)
                response = openai_client.chat.completions.create(
                    model="gpt-4o",
                    messages=messages,
                    max_tokens=4000,
                )
                result = response.choices[0].message.content
                st.session_state["compliance_result"] = result
                st.session_state["compliance_title"] = bid_title
                st.session_state["compliance_customer"] = customer

        if "compliance_result" in st.session_state:
            st.markdown("---")
            st.subheader("📊 Compliance Analysis Result")
            st.markdown(st.session_state["compliance_result"])

            st.markdown("---")
            col_save, col_clear = st.columns(2)
            with col_save:
                if st.button("💾 Save Analysis", type="primary", use_container_width=True):
                    insert_bid_analysis({
                        "org_id": org_id,
                        "created_by": user_email,
                        "bid_title": st.session_state["compliance_title"],
                        "customer": st.session_state["compliance_customer"],
                        "result": st.session_state["compliance_result"],
                        "analysis_date": str(date.today()),
                    })
                    st.success("✅ Analysis saved successfully!")
                    del st.session_state["compliance_result"]
                    del st.session_state["compliance_title"]
                    del st.session_state["compliance_customer"]
                    st.rerun()
            with col_clear:
                if st.button("🗑️ Discard", use_container_width=True):
                    del st.session_state["compliance_result"]
                    del st.session_state["compliance_title"]
                    del st.session_state["compliance_customer"]
                    st.rerun()

    # ── TAB 2: SAVED ANALYSES ──
    with tab2:
        st.subheader("📂 Saved Bid Analyses")

        if st.button("🔄 Load Records"):
            st.session_state["bid_analyses"] = get_bid_analyses(org_id)

        records = st.session_state.get("bid_analyses", [])

        if records:
            st.markdown(f"**{len(records)} record(s) found**")
            for r in records:
                label = f"📋 {r.get('analysis_date','')} — {r.get('bid_title','Untitled')} | 🏢 {r.get('customer','')}"
                with st.expander(label):
                    st.caption(f"👤 Added by: {r.get('created_by','')} on {r.get('analysis_date','')}")
                    st.markdown(r.get("result", ""))
                    if st.button("🗑️ Delete", key=f"del_bid_{r['id']}"):
                        delete_bid_analysis(r["id"])
                        st.session_state["bid_analyses"] = [x for x in records if x["id"] != r["id"]]
                        st.rerun()
        elif "bid_analyses" in st.session_state:
            st.info("No saved analyses yet.")
        else:
            st.info("Click 'Load Records' to view saved analyses.")
